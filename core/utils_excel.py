# Archivo: core/utils_excel.py
from django.apps import apps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import datetime

AZUL_OSCURO    = "1E3A5F"
AZUL_MEDIO     = "2E6DA4"
AZUL_CLARO     = "BDD7EE"
AZUL_MUY_CLARO = "DEEAF1"
VERDE          = "70AD47"
VERDE_CLARO    = "E2EFDA"
BLANCO         = "FFFFFF"
GRIS_BORDE     = "BFBFBF"
AMARILLO       = "FFD966"

def _borde():
    s = Side(style='thin', color=GRIS_BORDE)
    return Border(left=s, right=s, top=s, bottom=s)

def _borde_blanco():
    s = Side(style='thin', color=BLANCO)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _alinear(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_cell(ws, fila, col, valor, bg=None, color_texto=BLANCO, alin='center'):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(bold=True, color=color_texto, size=10, name="Arial")
    c.fill = _fill(bg or AZUL_MEDIO)
    c.alignment = _alinear(alin, 'center')
    c.border = _borde_blanco()
    return c

def _data_cell(ws, fila, col, valor, bg=BLANCO, alin='left'):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = _font(size=10)
    c.fill = _fill(bg)
    c.alignment = _alinear(alin, 'center')
    c.border = _borde()
    return c

def _titulo(ws, texto, total_cols, bg=AZUL_OSCURO):
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws['A1']
    c.value = texto
    c.font = Font(bold=True, color=BLANCO, size=14, name="Arial")
    c.fill = _fill(bg)
    c.alignment = _alinear('center', 'center')
    ws.row_dimensions[1].height = 30

def _fila_vacia(ws, i, fila, num_cols, cols_centro=None):
    bg = BLANCO if i % 2 == 0 else AZUL_MUY_CLARO
    ws.row_dimensions[fila].height = 18
    for col in range(1, num_cols + 1):
        c = ws.cell(row=fila, column=col, value=i if col == 1 else '')
        c.fill = _fill(bg)
        c.border = _borde()
        c.font = _font(size=10, color="BFBFBF")
        c.alignment = _alinear('center', 'center')


# ═══════════════════════════════════════════════════════════
# GENERADOR PARA ADMIN — Excel completo con todas las hojas
# ═══════════════════════════════════════════════════════════
def generar_excel_completo(expresos_qs, institucion="Unidad Educativa Expreso Escolar"):
    wb = Workbook()
    wb.remove(wb.active)

    datos_expresos = []
    todos_estudiantes = []
    for exp in expresos_qs:
        estudiantes = list(
            exp.estudiantes.filter(estado='activo')
            .select_related('padre__usuario')
            .order_by('curso', 'paralelo', 'apellido')
        )
        datos_expresos.append({'exp': exp, 'estudiantes': estudiantes})
        todos_estudiantes.extend(estudiantes)

    anio = "2024 – 2025"

    # Hoja 1: Padrón General
    _hoja_padron(wb, todos_estudiantes, institucion)

    # Hoja 2+: Lista por expreso (una hoja por expreso)
    for d in datos_expresos:
        _hoja_lista_expreso(wb, d['exp'], d['estudiantes'], anio)

    # Hoja N+: Lista por salón de CADA expreso
    for d in datos_expresos:
        _hoja_salones_por_expreso(wb, d['exp'], d['estudiantes'], anio)

    # Última hoja: Resumen General
    _hoja_resumen(wb, datos_expresos)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.date.today()
    response['Content-Disposition'] = f'attachment; filename="Expreso_Escolar_{fecha}.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════
# GENERADOR PARA TRANSPORTISTA — Solo su expreso
# ═══════════════════════════════════════════════════════════
def generar_excel_transportista(expreso, institucion="Unidad Educativa Expreso Escolar"):
    wb = Workbook()
    wb.remove(wb.active)

    estudiantes = list(
        expreso.estudiantes.filter(estado='activo')
        .select_related('padre__usuario')
        .order_by('curso', 'paralelo', 'apellido')
    )
    anio = "2024 – 2025"

    # Hoja 1: Lista general del expreso
    _hoja_lista_expreso(wb, expreso, estudiantes, anio)

    # Hoja 2: Lista por salón de su expreso
    _hoja_salones_por_expreso(wb, expreso, estudiantes, anio)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.date.today()
    response['Content-Disposition'] = f'attachment; filename="MisEstudiantes_{expreso.placa}_{fecha}.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════
# HOJA: PADRÓN GENERAL
# ═══════════════════════════════════════════════════════════
def _hoja_padron(wb, estudiantes, institucion):
    ws = wb.create_sheet("Padrón General")
    ws.sheet_view.showGridLines = False

    anchos = [5, 10, 30, 12, 10, 22, 22, 15, 15, 12, 10]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, "📋  PADRÓN GENERAL DE ESTUDIANTES – EXPRESO ESCOLAR", 11)

    ws.merge_cells('A2:K2')
    c = ws['A2']
    c.value = institucion
    c.font = Font(bold=True, color=BLANCO, size=11, name="Arial")
    c.fill = _fill(AZUL_MEDIO)
    c.alignment = _alinear('center', 'center')
    ws.row_dimensions[2].height = 20

    headers = ['N°', 'CÓDIGO', 'APELLIDOS Y NOMBRES', 'CURSO', 'SECCIÓN',
               'DIRECCIÓN', 'REPRESENTANTE', 'TELÉFONO 1', 'TELÉFONO 2',
               'EXPRESO ASIG.', 'ESTADO']
    ws.row_dimensions[3].height = 22
    for i, h in enumerate(headers, 1):
        _header_cell(ws, 3, i, h)

    for idx, est in enumerate(estudiantes, 1):
        fila = idx + 3
        bg = BLANCO if idx % 2 == 0 else AZUL_MUY_CLARO
        ws.row_dimensions[fila].height = 18
        vals = [
            idx, f"EST{idx:03d}",
            f"{est.apellido}, {est.nombre}",
            est.curso, est.paralelo, est.direccion,
            est.padre.usuario.get_full_name(),
            est.padre.telefono, '',
            est.expreso.nombre if est.expreso else '—',
            est.get_estado_display()
        ]
        centrados = [1, 4, 5, 10, 11]
        for col, val in enumerate(vals, 1):
            _data_cell(ws, fila, col, val, bg, 'center' if col in centrados else 'left')

    for i in range(len(estudiantes) + 1, 51):
        _fila_vacia(ws, i, i + 3, 11)


# ═══════════════════════════════════════════════════════════
# HOJA: LISTA POR EXPRESO (una hoja por expreso)
# ═══════════════════════════════════════════════════════════
def _hoja_lista_expreso(wb, exp, estudiantes, anio):
    nombre_hoja = f"Lista {exp.placa}"[:31]
    ws = wb.create_sheet(nombre_hoja)
    ws.sheet_view.showGridLines = False

    anchos = [5, 10, 30, 16, 22, 22, 15, 10, 22]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, "🗂️  LISTA DE ESTUDIANTES POR EXPRESO", 9)

    trans_nombre = exp.transportista.get_full_name() if exp.transportista else '—'
    trans_tel    = exp.transportista.telefono if exp.transportista and exp.transportista.telefono else '—'

    info = [
        ('EXPRESO:',      exp.nombre,    'AÑO LECTIVO:', anio),
        ('TRANSPORTISTA:', trans_nombre, 'PLACA:',       exp.placa),
        ('TELÉFONO:',     trans_tel,     'CAPACIDAD:',   exp.capacidad),
    ]
    for fi, (l1, v1, l2, v2) in enumerate(info, 2):
        ws.row_dimensions[fi].height = 18
        ws.merge_cells(f'A{fi}:B{fi}')
        ws.merge_cells(f'C{fi}:E{fi}')
        ws.merge_cells(f'F{fi}:G{fi}')
        ws.merge_cells(f'H{fi}:I{fi}')
        for col_l, val, es_lab in [('A', l1, True), ('C', v1, False), ('F', l2, True), ('H', v2, False)]:
            c = ws[f'{col_l}{fi}']
            c.value = val
            c.font = Font(bold=es_lab, size=10, name="Arial",
                          color=AZUL_OSCURO if es_lab else "000000")
            c.fill = _fill(AZUL_CLARO if es_lab else BLANCO)
            c.border = _borde()
            c.alignment = _alinear('right' if es_lab else 'left', 'center')

    headers = ['N°', 'CÓDIGO', 'APELLIDOS Y NOMBRES', 'CURSO',
               'DIRECCIÓN', 'REPRESENTANTE', 'TELÉFONO', 'ESTADO', 'OBSERVACIONES']
    ws.row_dimensions[5].height = 22
    for i, h in enumerate(headers, 1):
        _header_cell(ws, 5, i, h)

    for idx, est in enumerate(estudiantes, 1):
        fila = idx + 5
        bg = BLANCO if idx % 2 == 0 else AZUL_MUY_CLARO
        ws.row_dimensions[fila].height = 18
        vals = [idx, f"EST{idx:03d}",
                f"{est.apellido}, {est.nombre}",
                f"{est.curso} {est.paralelo}",
                est.direccion,
                est.padre.usuario.get_full_name(),
                est.padre.telefono,
                est.get_estado_display(), '']
        centrados = [1, 4, 8]
        for col, val in enumerate(vals, 1):
            _data_cell(ws, fila, col, val, bg, 'center' if col in centrados else 'left')

    for i in range(len(estudiantes) + 1, 31):
        _fila_vacia(ws, i, i + 5, 9)

    fila_t = 36
    ws.row_dimensions[fila_t].height = 22
    ws.merge_cells(f'A{fila_t}:G{fila_t}')
    c = ws[f'A{fila_t}']
    c.value = "TOTAL ESTUDIANTES EN ESTE EXPRESO:"
    c.font = Font(bold=True, size=10, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_MEDIO)
    c.alignment = _alinear('right', 'center')
    c.border = _borde()
    ws.merge_cells(f'H{fila_t}:I{fila_t}')
    c = ws[f'H{fila_t}']
    c.value = len(estudiantes)
    c.font = Font(bold=True, size=14, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = _alinear('center', 'center')
    c.border = _borde()


# ═══════════════════════════════════════════════════════════
# HOJA: SALONES POR EXPRESO (agrupado por curso/paralelo)
# ═══════════════════════════════════════════════════════════
def _hoja_salones_por_expreso(wb, exp, estudiantes, anio):
    nombre_hoja = f"Salones {exp.placa}"[:31]
    ws = wb.create_sheet(nombre_hoja)
    ws.sheet_view.showGridLines = False

    anchos = [5, 10, 30, 22, 22, 15, 22]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, "🏫  LISTA DE ESTUDIANTES POR SALÓN – " + exp.nombre.upper(), 7)

    trans_nombre = exp.transportista.get_full_name() if exp.transportista else '—'
    trans_tel    = exp.transportista.telefono if exp.transportista and exp.transportista.telefono else '—'

    # Info expreso fila 2
    ws.row_dimensions[2].height = 18
    ws.merge_cells('A2:B2'); ws.merge_cells('C2:D2')
    ws.merge_cells('E2:F2'); ws.merge_cells('G2:G2')
    pares = [('A', 'EXPRESO:', True), ('C', exp.nombre, False),
             ('E', 'PLACA:', True),   ('G', exp.placa, False)]
    for col_l, val, es_lab in pares:
        c = ws[f'{col_l}2']
        c.value = val
        c.font = Font(bold=es_lab, size=10, name="Arial",
                      color=AZUL_OSCURO if es_lab else "000000")
        c.fill = _fill(AZUL_CLARO if es_lab else BLANCO)
        c.border = _borde()
        c.alignment = _alinear('right' if es_lab else 'left', 'center')

    ws.row_dimensions[3].height = 18
    ws.merge_cells('A3:B3'); ws.merge_cells('C3:D3')
    ws.merge_cells('E3:F3'); ws.merge_cells('G3:G3')
    pares3 = [('A', 'TRANSPORTISTA:', True), ('C', trans_nombre, False),
              ('E', 'TELÉFONO:', True),       ('G', trans_tel, False)]
    for col_l, val, es_lab in pares3:
        c = ws[f'{col_l}3']
        c.value = val
        c.font = Font(bold=es_lab, size=10, name="Arial",
                      color=AZUL_OSCURO if es_lab else "000000")
        c.fill = _fill(AZUL_CLARO if es_lab else BLANCO)
        c.border = _borde()
        c.alignment = _alinear('right' if es_lab else 'left', 'center')

    # Agrupar por curso + paralelo
    from itertools import groupby
    estudiantes_sorted = sorted(estudiantes, key=lambda e: (e.curso, e.paralelo, e.apellido))
    grupos = []
    for (curso, paralelo), grupo in groupby(estudiantes_sorted, key=lambda e: (e.curso, e.paralelo)):
        grupos.append({'curso': curso, 'paralelo': paralelo, 'estudiantes': list(grupo)})

    fila_actual = 5

    for grupo in grupos:
        nombre_salon = f"{grupo['curso']} \"{grupo['paralelo']}\""
        ests = grupo['estudiantes']

        # Encabezado del salón
        ws.row_dimensions[fila_actual].height = 24
        ws.merge_cells(f'A{fila_actual}:E{fila_actual}')
        c = ws[f'A{fila_actual}']
        c.value = f"🏫  SALÓN: {nombre_salon}"
        c.font = Font(bold=True, color=BLANCO, size=11, name="Arial")
        c.fill = _fill(AZUL_OSCURO)
        c.alignment = _alinear('left', 'center')
        c.border = _borde_blanco()

        ws.merge_cells(f'F{fila_actual}:G{fila_actual}')
        c = ws[f'F{fila_actual}']
        c.value = f"Total: {len(ests)} alumno{'s' if len(ests) != 1 else ''}"
        c.font = Font(bold=True, color=BLANCO, size=10, name="Arial")
        c.fill = _fill(AZUL_MEDIO)
        c.alignment = _alinear('center', 'center')
        c.border = _borde_blanco()
        fila_actual += 1

        # Encabezados columnas
        ws.row_dimensions[fila_actual].height = 20
        headers_s = ['N°', 'CÓDIGO', 'APELLIDOS Y NOMBRES',
                     'REPRESENTANTE', 'TELÉFONO', 'DIRECCIÓN', 'ESTADO']
        for i, h in enumerate(headers_s, 1):
            _header_cell(ws, fila_actual, i, h, bg=AZUL_MEDIO)
        fila_actual += 1

        # Filas de estudiantes
        for idx, est in enumerate(ests, 1):
            bg = BLANCO if idx % 2 == 0 else AZUL_MUY_CLARO
            ws.row_dimensions[fila_actual].height = 18
            vals = [idx, f"EST{idx:03d}",
                    f"{est.apellido}, {est.nombre}",
                    est.padre.usuario.get_full_name(),
                    est.padre.telefono,
                    est.direccion,
                    est.get_estado_display()]
            centrados = [1, 5, 7]
            for col, val in enumerate(vals, 1):
                _data_cell(ws, fila_actual, col, val, bg,
                           'center' if col in centrados else 'left')
            fila_actual += 1

        # Fila de subtotal del salón
        ws.row_dimensions[fila_actual].height = 18
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        c = ws[f'A{fila_actual}']
        c.value = f"Subtotal {nombre_salon}:"
        c.font = Font(bold=True, size=10, name="Arial", color=AZUL_OSCURO)
        c.fill = _fill(AMARILLO)
        c.alignment = _alinear('right', 'center')
        c.border = _borde()
        c = ws.cell(row=fila_actual, column=7, value=len(ests))
        c.font = Font(bold=True, size=11, name="Arial", color=AZUL_OSCURO)
        c.fill = _fill(AMARILLO)
        c.alignment = _alinear('center', 'center')
        c.border = _borde()
        fila_actual += 2  # espacio entre salones

    # Total general al final
    ws.row_dimensions[fila_actual].height = 22
    ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
    c = ws[f'A{fila_actual}']
    c.value = "TOTAL GENERAL DE ESTUDIANTES:"
    c.font = Font(bold=True, size=11, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = _alinear('right', 'center')
    c.border = _borde()
    c = ws.cell(row=fila_actual, column=7, value=len(estudiantes))
    c.font = Font(bold=True, size=14, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_OSCURO)
    c.alignment = _alinear('center', 'center')
    c.border = _borde()


# ═══════════════════════════════════════════════════════════
# HOJA: RESUMEN GENERAL
# ═══════════════════════════════════════════════════════════
def _hoja_resumen(wb, datos_expresos):
    ws = wb.create_sheet("Resumen General")
    ws.sheet_view.showGridLines = False
    for i, a in enumerate([30, 25, 20, 18, 12, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, "📊  RESUMEN GENERAL – EXPRESO ESCOLAR", 7)

    total = sum(len(d['estudiantes']) for d in datos_expresos)
    stats = [
        ("TOTAL ESTUDIANTES REGISTRADOS", total),
        ("ESTUDIANTES ACTIVOS", total),
        ("ESTUDIANTES INACTIVOS", 0),
        ("EXPRESOS EN OPERACIÓN", len(datos_expresos)),
    ]
    for fi, (label, valor) in enumerate(stats, 2):
        ws.row_dimensions[fi].height = 22
        ws.merge_cells(f'A{fi}:E{fi}')
        c = ws[f'A{fi}']
        c.value = label
        c.font = Font(bold=True, size=11, name="Arial", color=AZUL_OSCURO)
        c.fill = _fill(AZUL_CLARO if fi % 2 == 0 else AZUL_MUY_CLARO)
        c.alignment = _alinear('left', 'center')
        c.border = _borde()
        ws.merge_cells(f'F{fi}:G{fi}')
        c = ws[f'F{fi}']
        c.value = valor
        c.font = Font(bold=True, size=14, name="Arial", color=BLANCO)
        c.fill = _fill(AZUL_MEDIO)
        c.alignment = _alinear('center', 'center')
        c.border = _borde()

    ws.row_dimensions[7].height = 8
    headers = ['EXPRESO', 'TRANSPORTISTA', 'PLACA', 'TELÉFONO', 'CAPACIDAD', 'ESTUDIANTES', 'ESTADO']
    ws.row_dimensions[8].height = 22
    for i, h in enumerate(headers, 1):
        _header_cell(ws, 8, i, h)

    for fi, d in enumerate(datos_expresos, 9):
        exp = d['exp']
        bg = BLANCO if fi % 2 == 0 else AZUL_MUY_CLARO
        ws.row_dimensions[fi].height = 18
        trans = exp.transportista
        vals = [
            exp.nombre,
            trans.get_full_name() if trans else '—',
            exp.placa,
            trans.telefono if trans and trans.telefono else '—',
            exp.capacidad,
            len(d['estudiantes']),
            'Activo' if exp.activo else 'Inactivo'
        ]
        centrados = [3, 5, 6, 7]
        for col, val in enumerate(vals, 1):
            _data_cell(ws, fi, col, val, bg, 'center' if col in centrados else 'left')

# ═══════════════════════════════════════════════════════════
# GENERADOR PARA UN SOLO EXPRESO — Lista + Salones
# ═══════════════════════════════════════════════════════════
def generar_excel_un_expreso(expreso, institucion="Unidad Educativa Expreso Escolar"):
    wb = Workbook()
    wb.remove(wb.active)

    estudiantes = list(
        expreso.estudiantes.filter(estado='activo')
        .select_related('padre__usuario')
        .order_by('curso', 'paralelo', 'apellido')
    )
    anio = "2024 – 2025"

    # Hoja 1: Lista de estudiantes por expreso
    _hoja_lista_expreso(wb, expreso, estudiantes, anio)

    # Hoja 2: Lista por salones
    _hoja_salones_por_expreso(wb, expreso, estudiantes, anio)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.date.today()
    response['Content-Disposition'] = f'attachment; filename="Expreso_{expreso.placa}_{fecha}.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════
# GENERADOR SOLO SALONES — Para el tab de salones
# ═══════════════════════════════════════════════════════════
def generar_excel_salones_expreso(expreso, institucion="Unidad Educativa Expreso Escolar"):
    wb = Workbook()
    wb.remove(wb.active)

    estudiantes = list(
        expreso.estudiantes.filter(estado='activo')
        .select_related('padre__usuario')
        .order_by('curso', 'paralelo', 'apellido')
    )

    _hoja_salones_por_expreso(wb, expreso, estudiantes, "2024 – 2025")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.date.today()
    response['Content-Disposition'] = f'attachment; filename="Salones_{expreso.placa}_{fecha}.xlsx"'
    wb.save(response)
    return response

def generar_excel_un_salon(expreso, curso, paralelo, institucion="Unidad Educativa Expreso Escolar"):
    wb = Workbook()
    wb.remove(wb.active)

    Estudiante = apps.get_model('core', 'Estudiante')
    estudiantes = list(
        Estudiante.objects.filter(
            expreso=expreso,
            estado='activo',
            curso=curso,
            paralelo=paralelo
        ).select_related('padre__usuario')
        .order_by('apellido')
    )

    ws = wb.create_sheet(f"{curso} {paralelo}"[:31])
    ws.sheet_view.showGridLines = False

    anchos = [5, 10, 30, 22, 15, 22, 12]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, f"LISTA DE ESTUDIANTES - {curso} '{paralelo}' - {expreso.nombre.upper()}", 7)

    trans_nombre = expreso.transportista.get_full_name() if expreso.transportista else '—'
    trans_tel    = expreso.transportista.telefono if expreso.transportista and expreso.transportista.telefono else '—'

    ws.row_dimensions[2].height = 18
    ws.merge_cells('A2:C2'); ws.merge_cells('D2:E2')
    ws.merge_cells('F2:F2'); ws.merge_cells('G2:G2')
    for col_l, val, es_lab in [('A','EXPRESO:',True),('D',expreso.nombre,False),('F','PLACA:',True),('G',expreso.placa,False)]:
        c = ws[f'{col_l}2']; c.value = val
        c.font = Font(bold=es_lab, size=10, name="Arial", color=AZUL_OSCURO if es_lab else "000000")
        c.fill = _fill(AZUL_CLARO if es_lab else BLANCO); c.border = _borde(); c.alignment = _alinear('right' if es_lab else 'left','center')

    ws.row_dimensions[3].height = 18
    ws.merge_cells('A3:C3'); ws.merge_cells('D3:E3')
    ws.merge_cells('F3:F3'); ws.merge_cells('G3:G3')
    for col_l, val, es_lab in [('A','TRANSPORTISTA:',True),('D',trans_nombre,False),('F','TELEFONO:',True),('G',trans_tel,False)]:
        c = ws[f'{col_l}3']; c.value = val
        c.font = Font(bold=es_lab, size=10, name="Arial", color=AZUL_OSCURO if es_lab else "000000")
        c.fill = _fill(AZUL_CLARO if es_lab else BLANCO); c.border = _borde(); c.alignment = _alinear('right' if es_lab else 'left','center')

    headers = ['N°', 'CÓDIGO', 'APELLIDOS Y NOMBRES', 'REPRESENTANTE', 'TELÉFONO', 'DIRECCIÓN', 'ESTADO']
    ws.row_dimensions[4].height = 22
    for i, h in enumerate(headers, 1):
        _header_cell(ws, 4, i, h)

    for idx, est in enumerate(estudiantes, 1):
        fila = idx + 4
        bg = BLANCO if idx % 2 == 0 else AZUL_MUY_CLARO
        ws.row_dimensions[fila].height = 18
        vals = [idx, f"EST{idx:03d}",
                f"{est.apellido}, {est.nombre}",
                est.padre.usuario.get_full_name(),
                est.padre.telefono,
                est.direccion,
                est.get_estado_display()]
        for col, val in enumerate(vals, 1):
            _data_cell(ws, fila, col, val, bg, 'center' if col in [1, 7] else 'left')

    fila_t = len(estudiantes) + 6
    ws.merge_cells(f'A{fila_t}:F{fila_t}')
    c = ws[f'A{fila_t}']
    c.value = f"TOTAL ESTUDIANTES - {curso} '{paralelo}':"
    c.font = Font(bold=True, size=10, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_MEDIO); c.alignment = _alinear('right','center'); c.border = _borde()
    c = ws.cell(row=fila_t, column=7, value=len(estudiantes))
    c.font = Font(bold=True, size=14, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_OSCURO); c.alignment = _alinear('center','center'); c.border = _borde()
    ws.row_dimensions[fila_t].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Salon_{curso}_{paralelo}_{expreso.placa}".replace(' ', '_').replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response

def generar_excel_salon_profesor(curso, paralelo):
    from django.apps import apps
    Estudiante = apps.get_model('core', 'Estudiante')

    wb = Workbook()
    wb.remove(wb.active)

    estudiantes = list(
        Estudiante.objects.filter(
            estado='activo',
            curso=curso,
            paralelo=paralelo
        ).select_related('expreso__transportista', 'padre__usuario')
         .order_by('apellido')
    )

    ws = wb.create_sheet(f"{curso} {paralelo}"[:31])
    ws.sheet_view.showGridLines = False

    anchos = [5, 18, 18, 14, 20, 16, 22, 16]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    _titulo(ws, f"LISTADO POR SALON - {curso} '{paralelo}'", 8)

    headers = ['N°', 'APELLIDO', 'NOMBRE', 'EXPRESO',
               'TRANSPORTISTA', 'TEL. TRANSPORTISTA',
               'REPRESENTANTE', 'TEL. REPRESENTANTE']
    ws.row_dimensions[3].height = 22
    for i, h in enumerate(headers, 1):
        _header_cell(ws, 3, i, h)

    for idx, est in enumerate(estudiantes, 1):
        fila = idx + 3
        bg = BLANCO if idx % 2 == 0 else AZUL_MUY_CLARO
        ws.row_dimensions[fila].height = 18

        expreso_nombre = est.expreso.nombre if est.expreso else '—'
        trans_nombre = est.expreso.transportista.get_full_name() if est.expreso and est.expreso.transportista else '—'
        trans_tel = est.expreso.transportista.telefono if est.expreso and est.expreso.transportista else '—'

        vals = [idx, est.apellido, est.nombre, expreso_nombre,
                trans_nombre, trans_tel,
                est.padre.usuario.get_full_name(), est.padre.telefono]

        for col, val in enumerate(vals, 1):
            _data_cell(ws, fila, col, val, bg, 'center' if col in [1] else 'left')

    fila_t = len(estudiantes) + 5
    ws.merge_cells(f'A{fila_t}:G{fila_t}')
    c = ws[f'A{fila_t}']
    c.value = f"TOTAL ESTUDIANTES - {curso} '{paralelo}':"
    c.font = Font(bold=True, size=10, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_MEDIO); c.alignment = _alinear('right','center'); c.border = _borde()
    c = ws.cell(row=fila_t, column=8, value=len(estudiantes))
    c.font = Font(bold=True, size=14, name="Arial", color=BLANCO)
    c.fill = _fill(AZUL_OSCURO); c.alignment = _alinear('center','center'); c.border = _borde()
    ws.row_dimensions[fila_t].height = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Salon_{curso}_{paralelo}".replace(' ', '_').replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    wb.save(response)
    return response